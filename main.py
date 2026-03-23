#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Shopline 進貨單自動化工具 v3.0
使用方式：python3 shopline進貨單工具.py
需要：Python 3.8+
"""

import sys, os, json, threading, re, io, tempfile, uuid, time
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import parse_qs, urlparse

# ── 自動安裝套件 ──────────────────────────────────────────────────────────────
def _pip_install(pkg):
    import subprocess
    print(f"  安裝 {pkg} 中...")
    for cmd in [
        [sys.executable, "-m", "pip", "install", pkg, "--quiet", "--user"],
        [sys.executable, "-m", "pip", "install", pkg, "--quiet"],
        [sys.executable, "-m", "pip", "install", pkg, "--quiet", "--break-system-packages"],
        ["pip3", "install", pkg, "--quiet", "--user"],
    ]:
        try:
            if subprocess.call(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL) == 0:
                return True
        except Exception:
            continue
    return False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    _pip_install("openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    for pkg in ["xlrd==1.2.0", "xlrd"]:
        if _pip_install(pkg):
            try:
                import xlrd
                HAS_XLRD = True
                break
            except ImportError:
                pass
    else:
        HAS_XLRD = False

# ── HTML 介面 ─────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>進貨單自動化工具</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Noto+Sans+TC:wght@300;400;500;700&display=swap');
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --ink:#1a1a1a;--ink2:#444;--ink3:#888;--line:#e0e0e0;--line2:#f0f0ee;
  --bg:#fafaf8;--white:#fff;--ok:#2a7a4b;--ok-bg:#edf7f1;
  --warn:#a85c00;--warn-bg:#fff8ed;--err:#b83232;--err-bg:#fff0f0;
  --tag:#f0f0ee;--r:10px;--mono:'DM Mono',monospace;--sans:'Noto Sans TC',sans-serif;
}
html{font-size:15px}
body{font-family:var(--sans);background:var(--bg);color:var(--ink);min-height:100vh;padding:0 0 80px}
.header{background:var(--white);border-bottom:1px solid var(--line);padding:18px 40px;display:flex;align-items:center;gap:14px;position:sticky;top:0;z-index:100}
.hicon{width:36px;height:36px;background:var(--ink);border-radius:8px;display:grid;place-items:center;font-size:18px}
.header h1{font-size:17px;font-weight:700;letter-spacing:-.3px}
.header p{font-size:12px;color:var(--ink3);margin-top:1px}
.vtag{margin-left:auto;font-family:var(--mono);font-size:11px;background:var(--tag);color:var(--ink3);padding:3px 9px;border-radius:20px}
.main{max-width:860px;margin:0 auto;padding:36px 24px 0}
.steps{display:flex;background:var(--white);border:1px solid var(--line);border-radius:var(--r);overflow:hidden;margin-bottom:28px}
.step{flex:1;padding:14px 16px;display:flex;align-items:center;gap:10px;border-right:1px solid var(--line);font-size:13px;color:var(--ink3)}
.step:last-child{border-right:none}
.step.active{background:var(--ink);color:#fff}
.step.done{background:var(--ok-bg);color:var(--ok)}
.snum{width:22px;height:22px;border-radius:50%;border:1.5px solid currentColor;display:grid;place-items:center;font-family:var(--mono);font-size:11px;flex-shrink:0}
.step.active .snum{background:#fff;color:var(--ink);border-color:#fff}
.step.done .snum{background:var(--ok);color:#fff;border-color:var(--ok)}
.card{background:var(--white);border:1px solid var(--line);border-radius:var(--r);margin-bottom:18px;overflow:hidden}
.ch{padding:15px 20px;border-bottom:1px solid var(--line2);display:flex;align-items:center;gap:10px}
.ch h2{font-size:14px;font-weight:600}
.badge{font-family:var(--mono);font-size:10px;padding:2px 8px;border-radius:20px;background:var(--tag);color:var(--ink3)}
.badge.req{background:#fff3f3;color:var(--err)}
.cb{padding:20px}
.ugrid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.uzone{border:1.5px dashed var(--line);border-radius:8px;padding:22px 16px;text-align:center;cursor:pointer;transition:all .15s;background:var(--bg)}
.uzone:hover{border-color:var(--ink);background:#f5f5f3}
.uzone.loaded{border-color:var(--ok);border-style:solid;background:var(--ok-bg)}
.uzone input{display:none}
.uicon{font-size:26px;margin-bottom:8px}
.ulabel{font-size:13px;font-weight:500;color:var(--ink2);margin-bottom:4px}
.uhint{font-size:11px;color:var(--ink3);font-family:var(--mono)}
.ufn{font-size:12px;font-family:var(--mono);color:var(--ok);margin-top:6px;font-weight:500;word-break:break-all}
.sgrid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.field label{display:block;font-size:12px;font-weight:500;color:var(--ink3);margin-bottom:6px;text-transform:uppercase;letter-spacing:.05em}
select{width:100%;height:42px;padding:0 14px;border:1px solid var(--line);border-radius:7px;font-size:14px;font-family:var(--sans);background:var(--white);color:var(--ink);appearance:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%23888' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 14px center}
select:focus{outline:none;border-color:var(--ink)}
select:disabled{opacity:.4;cursor:not-allowed}
.rbtn{width:100%;height:52px;background:var(--ink);color:#fff;border:none;border-radius:8px;font-size:15px;font-weight:600;font-family:var(--sans);cursor:pointer;display:flex;align-items:center;justify-content:center;gap:8px;transition:opacity .15s}
.rbtn:hover:not(:disabled){opacity:.88}
.rbtn:disabled{opacity:.35;cursor:not-allowed}
.spin{width:18px;height:18px;border:2px solid rgba(255,255,255,.35);border-top-color:#fff;border-radius:50%;animation:spin .7s linear infinite;display:none}
.rbtn.loading .spin{display:block}
@keyframes spin{to{transform:rotate(360deg)}}
#results{display:none}
.rbanner{border-radius:var(--r);padding:16px 20px;margin-bottom:16px;display:flex;align-items:center;gap:12px}
.rbanner.ok{background:var(--ok-bg);border:1px solid #b6ddc7}
.rbanner.warn{background:var(--warn-bg);border:1px solid #f5d89a}
.bicon{font-size:22px}
.btitle{font-weight:600;font-size:14px}
.bsub{font-size:13px;color:var(--ink2);margin-top:2px}
.dlgrid{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:18px}
.dlbtn{display:flex;align-items:center;gap:12px;padding:14px 18px;border-radius:8px;cursor:pointer;border:1px solid;font-family:var(--sans);font-size:14px;font-weight:500;transition:opacity .15s;text-decoration:none}
.dlbtn:hover{opacity:.8}
.dlbtn.pri{background:var(--ink);color:#fff;border-color:var(--ink)}
.dlbtn.sec{background:var(--white);color:var(--warn);border-color:#f5d89a}
.dlbtn.sec.hide{opacity:.3;pointer-events:none}
.dlicon{font-size:22px;flex-shrink:0}
.dllabel{font-size:12px;opacity:.7;margin-bottom:1px}
.dlname{font-size:14px;font-weight:600}
.stats{display:flex;gap:12px;margin-bottom:16px}
.sbox{flex:1;padding:12px 16px;border-radius:8px;border:1px solid var(--line);background:var(--white)}
.snum{font-family:var(--mono);font-size:26px;font-weight:500;line-height:1}
.slabel{font-size:12px;color:var(--ink3);margin-top:4px}
.sbox.ok .snum{color:var(--ok)}
.sbox.warn .snum{color:var(--warn)}
.alert{padding:12px 16px;border-radius:7px;font-size:13px;margin-bottom:16px;display:none}
.alert.err{background:var(--err-bg);color:var(--err);border:1px solid #f5b8b8}
.alert.show{display:block}
table.ut{width:100%;border-collapse:collapse;font-size:13px}
table.ut th{text-align:left;padding:8px 12px;background:var(--tag);font-size:11px;font-weight:600;color:var(--ink3);text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid var(--line)}
table.ut td{padding:10px 12px;border-bottom:1px solid var(--line2);vertical-align:top}
table.ut tr:last-child td{border-bottom:none}
.chip{font-family:var(--mono);font-size:12px;background:var(--tag);padding:2px 7px;border-radius:4px}
.wchip{font-size:11px;font-weight:600;padding:2px 8px;border-radius:4px;background:var(--warn-bg);color:var(--warn)}
.avail{font-size:11px;color:var(--ink3);font-family:var(--mono);line-height:1.7;margin-top:4px;white-space:pre-line}
.reset{display:block;margin:20px auto 0;background:none;border:none;color:var(--ink3);font-size:13px;font-family:var(--sans);cursor:pointer;text-decoration:underline}
</style>
</head>
<body>
<div class="header">
  <div class="hicon">📦</div>
  <div><h1>Shopline 進貨單自動化</h1><p>上傳進貨單與商品檔，自動產出可匯入的進貨單</p></div>
  <span class="vtag">v3.0 本機版</span>
</div>
<div class="main">
  <div class="steps">
    <div class="step active" id="step1"><div class="snum" id="sn1">1</div>上傳檔案</div>
    <div class="step" id="step2"><div class="snum" id="sn2">2</div>選擇分店 / 員工</div>
    <div class="step" id="step3"><div class="snum" id="sn3">3</div>產出進貨單</div>
  </div>
  <div class="alert err" id="alertBox"></div>
  <div class="card">
    <div class="ch"><h2>上傳檔案</h2><span class="badge req">必填</span></div>
    <div class="cb">
      <div class="ugrid">
        <div class="uzone" id="zone1" onclick="document.getElementById('f1').click()"
             ondragover="ev.preventDefault();this.style.borderColor='#1a1a1a'" 
             ondragleave="this.style.borderColor=''" 
             ondrop="ev.preventDefault();this.style.borderColor='';handleDrop(ev,true)">
          <input type="file" id="f1" accept=".xls,.xlsx" onchange="handleFile(this,true)">
          <div class="uicon">📋</div>
          <div class="ulabel">進貨單</div>
          <div class="uhint">.xls / .xlsx</div>
          <div class="ufn" id="fn1"></div>
        </div>
        <div class="uzone" id="zone2" onclick="document.getElementById('f2').click()"
             ondragover="ev.preventDefault();this.style.borderColor='#1a1a1a'"
             ondragleave="this.style.borderColor=''"
             ondrop="ev.preventDefault();this.style.borderColor='';handleDrop(ev,false)">
          <input type="file" id="f2" accept=".xlsx" onchange="handleFile(this,false)">
          <div class="uicon">🗂️</div>
          <div class="ulabel">總商品檔（shopline格式表）</div>
          <div class="uhint">.xlsx</div>
          <div class="ufn" id="fn2"></div>
        </div>
      </div>
    </div>
  </div>
  <div class="card">
    <div class="ch"><h2>選擇分店與進貨人員</h2><span class="badge req">必填</span></div>
    <div class="cb">
      <div class="sgrid">
        <div class="field"><label>進貨分店</label>
          <select id="storeSelect" disabled onchange="checkReady()"><option value="">── 請先上傳商品檔 ──</option></select>
        </div>
        <div class="field"><label>進貨人員</label>
          <select id="staffSelect" disabled onchange="checkReady()"><option value="">── 請先上傳商品檔 ──</option></select>
        </div>
      </div>
    </div>
  </div>
  <button class="rbtn" id="runBtn" disabled onclick="runProcess()">
    <div class="spin"></div><span>▶ 產出進貨單</span>
  </button>
  <div id="results">
    <div class="rbanner" id="rbanner"><div class="bicon" id="bicon"></div><div><div class="btitle" id="btitle"></div><div class="bsub" id="bsub"></div></div></div>
    <div class="stats">
      <div class="sbox ok"><div class="snum" id="sok">0</div><div class="slabel">成功比對</div></div>
      <div class="sbox warn"><div class="snum" id="swarn">0</div><div class="slabel">待確認</div></div>
      <div class="sbox"><div class="snum" id="stotal">0</div><div class="slabel">規格總計</div></div>
    </div>
    <div class="dlgrid">
      <a class="dlbtn pri" id="dlMain" href="#"><span class="dlicon">📥</span><div><div class="dllabel">下載並上傳至 Shopline</div><div class="dlname">進貨單_已填寫.xlsx</div></div></a>
      <a class="dlbtn sec hide" id="dlWarn" href="#"><span class="dlicon">⚠️</span><div><div class="dllabel">需要人工確認</div><div class="dlname">待確認清單.xlsx</div></div></a>
    </div>
    <div class="card" id="unmCard" style="display:none">
      <div class="ch"><h2>⚠️ 待確認項目</h2><span class="badge">需人工補齊</span></div>
      <div class="cb">
        <p style="font-size:13px;color:var(--ink2);margin-bottom:14px">以下規格在商品檔中找不到對應，請對照「現有規格」確認後，手動補填進貨單。</p>
        <table class="ut"><thead><tr><th>商品編號</th><th>商品名稱</th><th>進貨單規格</th><th>數量</th><th>問題</th><th>Shopline 現有規格</th></tr></thead>
        <tbody id="utbody"></tbody></table>
      </div>
    </div>
    <button class="reset" onclick="resetAll()">↩ 重新上傳</button>
  </div>
</div>
<script>
var f1Session=null, f2Session=null;

function handleDrop(ev, isOrder) {
  var f = ev.dataTransfer.files[0];
  if (f) doUpload(f, isOrder);
}
function handleFile(input, isOrder) {
  var f = input.files[0];
  if (f) doUpload(f, isOrder);
}

function doUpload(file, isOrder) {
  showAlert('');
  var fd = new FormData();
  fd.append('file', file);
  fd.append('type', isOrder ? 'order' : 'product');
  
  var xhr = new XMLHttpRequest();
  xhr.open('POST', '/upload', true);
  xhr.onreadystatechange = function() {
    if (xhr.readyState !== 4) return;
    try {
      var d = JSON.parse(xhr.responseText);
      if (d.error) { showAlert(d.error); return; }
      if (isOrder) {
        f1Session = d.session;
        document.getElementById('zone1').classList.add('loaded');
        document.getElementById('fn1').textContent = '✓ ' + file.name;
      } else {
        f2Session = d.session;
        document.getElementById('zone2').classList.add('loaded');
        document.getElementById('fn2').textContent = '✓ ' + file.name;
        populateSelects(d.channels, d.employees);
      }
      updateSteps();
      checkReady();
    } catch(e) {
      showAlert('上傳失敗：' + (xhr.responseText || e.message));
    }
  };
  xhr.onerror = function() { showAlert('網路錯誤，請確認工具是否正在運行'); };
  xhr.send(fd);
}

function populateSelects(channels, employees) {
  var ss = document.getElementById('storeSelect');
  var es = document.getElementById('staffSelect');
  ss.innerHTML = '<option value="">── 請選擇 ──</option>';
  es.innerHTML = '<option value="">── 請選擇 ──</option>';
  channels.forEach(function(c) {
    var o = document.createElement('option');
    o.value = JSON.stringify(c); o.textContent = c.name; ss.appendChild(o);
  });
  employees.forEach(function(e) {
    var o = document.createElement('option');
    o.value = JSON.stringify(e); o.textContent = e.name; es.appendChild(o);
  });
  ss.disabled = false; es.disabled = false;
}

function updateSteps() {
  var h1 = !!f1Session, h2 = !!f2Session;
  document.getElementById('step1').className = 'step ' + (h1&&h2 ? 'done' : 'active');
  document.getElementById('sn1').textContent = h1&&h2 ? '✓' : '1';
  document.getElementById('step2').className = 'step ' + (h1&&h2 ? 'active' : '');
}

function checkReady() {
  var s = document.getElementById('storeSelect').value;
  var st = document.getElementById('staffSelect').value;
  var ok = f1Session && f2Session && s && st;
  document.getElementById('runBtn').disabled = !ok;
  if (ok) {
    document.getElementById('step2').className = 'step done';
    document.getElementById('sn2').textContent = '✓';
    document.getElementById('step3').className = 'step active';
  }
}

function runProcess() {
  var btn = document.getElementById('runBtn');
  btn.classList.add('loading'); btn.disabled = true;
  var store = JSON.parse(document.getElementById('storeSelect').value);
  var staff = JSON.parse(document.getElementById('staffSelect').value);
  
  var xhr = new XMLHttpRequest();
  xhr.open('POST', '/process', true);
  xhr.setRequestHeader('Content-Type', 'application/json');
  xhr.onreadystatechange = function() {
    if (xhr.readyState !== 4) return;
    btn.classList.remove('loading'); btn.disabled = false;
    try {
      var d = JSON.parse(xhr.responseText);
      if (d.error) { showAlert(d.error); return; }
      showResults(d);
      document.getElementById('step3').className = 'step done';
      document.getElementById('sn3').textContent = '✓';
    } catch(e) {
      showAlert('處理失敗：' + (xhr.responseText || e.message));
    }
  };
  xhr.onerror = function() { 
    btn.classList.remove('loading'); btn.disabled = false;
    showAlert('連線失敗，請確認工具正在運行');
  };
  xhr.send(JSON.stringify({
    order_session: f1Session,
    product_session: f2Session,
    store: store,
    staff: staff
  }));
}

function showResults(d) {
  document.getElementById('sok').textContent = d.matched;
  document.getElementById('swarn').textContent = d.unmatched.length;
  document.getElementById('stotal').textContent = d.total;
  var rb = document.getElementById('rbanner');
  if (!d.unmatched.length) {
    rb.className = 'rbanner ok';
    document.getElementById('bicon').textContent = '✅';
    document.getElementById('btitle').textContent = '全部比對成功！';
    document.getElementById('bsub').textContent = '共 ' + d.matched + ' 筆，可直接下載並上傳至 Shopline。';
  } else {
    rb.className = 'rbanner warn';
    document.getElementById('bicon').textContent = '⚠️';
    document.getElementById('btitle').textContent = d.matched + ' 筆成功，' + d.unmatched.length + ' 筆需要確認';
    document.getElementById('bsub').textContent = '請下載進貨單後，再對照待確認清單手動補齊。';
  }
  document.getElementById('dlMain').href = '/download?session=' + d.session + '&type=main';
  document.getElementById('dlMain').setAttribute('download', '進貨單_已填寫.xlsx');
  if (d.unmatched.length) {
    document.getElementById('dlWarn').href = '/download?session=' + d.session + '&type=warn';
    document.getElementById('dlWarn').setAttribute('download', '待確認清單.xlsx');
    document.getElementById('dlWarn').className = 'dlbtn sec';
    document.getElementById('unmCard').style.display = '';
    var tbody = document.getElementById('utbody');
    tbody.innerHTML = '';
    d.unmatched.forEach(function(u) {
      var tr = document.createElement('tr');
      tr.innerHTML = '<td><span class="chip">' + u.code + '</span></td>' +
        '<td style="font-size:12px;max-width:180px">' + u.name + '</td>' +
        '<td><strong>' + u.color + '</strong> ' + u.size + '</td>' +
        '<td style="font-family:var(--mono)">' + u.qty + '</td>' +
        '<td><span class="wchip">' + u.issue + '</span></td>' +
        '<td><div class="avail">' + (u.avail||'').split('、').slice(0,8).join('\n') + '</div></td>';
      tbody.appendChild(tr);
    });
  }
  document.getElementById('results').style.display = '';
  document.getElementById('results').scrollIntoView({behavior:'smooth'});
}

function resetAll() {
  f1Session = f2Session = null;
  ['zone1','zone2'].forEach(function(z){document.getElementById(z).classList.remove('loaded')});
  ['fn1','fn2'].forEach(function(f){document.getElementById(f).textContent=''});
  ['f1','f2'].forEach(function(f){document.getElementById(f).value=''});
  document.getElementById('storeSelect').innerHTML='<option value="">── 請先上傳商品檔 ──</option>';
  document.getElementById('storeSelect').disabled=true;
  document.getElementById('staffSelect').innerHTML='<option value="">── 請先上傳商品檔 ──</option>';
  document.getElementById('staffSelect').disabled=true;
  document.getElementById('runBtn').disabled=true;
  document.getElementById('results').style.display='none';
  document.getElementById('unmCard').style.display='none';
  document.getElementById('dlWarn').className='dlbtn sec hide';
  showAlert('');
  ['step1','step2','step3'].forEach(function(s,i){
    document.getElementById(s).className='step'+(i===0?' active':'');
    document.getElementById('sn'+(i+1)).textContent=i+1;
  });
  window.scrollTo({top:0,behavior:'smooth'});
}

function showAlert(msg) {
  var el = document.getElementById('alertBox');
  el.textContent = msg;
  el.className = 'alert err' + (msg ? ' show' : '');
}
</script>
</body>
</html>"""

# ── Matching logic ────────────────────────────────────────────────────────────
def find_product(code, name, products):
    slash_parts = sorted(code.split('/'), key=len, reverse=True)
    for part in slash_parts:
        if len(part) < 3: continue
        m = [p for p in products if part.lower() in p['productName'].lower()]
        if m: return m
    m = [p for p in products if code.lower() in p['productName'].lower()]
    if m: return m
    clean = re.sub(r'\[.*?\]', '', name).strip()
    for kl in [10, 8, 6]:
        m = [p for p in products if clean[:kl].lower() in p['productName'].lower()]
        if m: return m
    for part in slash_parts:
        if len(part) < 3: continue
        m = [p for p in products if p['sku'].upper().startswith(part.upper())]
        if m: return m
    return []

def find_variation(prods, color, size):
    alt = {'黑': '⿊', '⿊': '黑'}
    for c in ([color, alt[color]] if color in alt else [color]):
        exact = f"{c} {size}"
        vm = [p for p in prods if p['variationCN'].strip() == exact]
        if vm: return vm[0]
        pat = re.compile(r'(?<![A-Za-z0-9])' + re.escape(size) + r'(?![A-Za-z0-9])')
        vm = [p for p in prods if c in p['variationCN'] and pat.search(p['variationCN'])]
        if vm: return vm[0]
        vm = [p for p in prods if p['variationCN'].strip() == c]
        if vm: return vm[0]
        vm = [p for p in prods if p['variationCN'].strip().lower() == c.lower()]
        if vm: return vm[0]
        vm = [p for p in prods if p['variationCN'].lower().startswith(c.lower())]
        if vm:
            if size:
                vm2 = [p for p in vm if size in p['variationCN']]
                if vm2: return vm2[0]
            return vm[0]
    return None

# ── File parsing ──────────────────────────────────────────────────────────────
def parse_order_bytes(data):
    rows = []
    # Detect format by magic bytes
    is_xlsx = data[:4] == b'PK\x03\x04'
    suffix = '.xlsx' if is_xlsx else '.xls'
    
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tf:
        tf.write(data)
        tmp = tf.name
    
    try:
        if is_xlsx:
            wb = openpyxl.load_workbook(tmp, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' 
                      for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row): continue
                r = dict(zip(headers, [str(v).strip() if v is not None else '' for v in row]))
                if r.get('商品編號'):
                    rows.append({'code': r.get('商品編號',''), 'name': r.get('商品名稱',''),
                                 'color': r.get('商品樣式',''), 'size': r.get('商品尺寸','')})
        else:
            if not HAS_XLRD:
                raise Exception('讀取 .xls 需要 xlrd 套件，請重新啟動工具（會自動安裝）')
            book = xlrd.open_workbook(tmp)
            sh = book.sheet_by_index(0)
            headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
            for rx in range(1, sh.nrows):
                r = {headers[c]: str(sh.cell_value(rx, c)).strip() for c in range(sh.ncols)}
                if r.get('商品編號'):
                    rows.append({'code': r.get('商品編號',''), 'name': r.get('商品名稱',''),
                                 'color': r.get('商品樣式',''), 'size': r.get('商品尺寸','')})
    finally:
        try: os.unlink(tmp)
        except: pass
    return rows

def parse_product_bytes(data):
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
        tf.write(data); tmp = tf.name
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        products = []
        ws = wb['ID_product and variation']
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            products.append({
                'productId':   str(row[0]).strip(),
                'productName': str(row[1]).strip() if row[1] else '',
                'variationId': str(row[3]).strip() if row[3] else '',
                'variationCN': str(row[4]).replace('\t',' ').strip() if row[4] else '',
                'sku':         str(row[6]).strip() if row[6] else '',
                'cost':        row[7] if row[7] is not None else '',
            })
        wc = wb['ID_channel']
        channels = []
        for row in wc.iter_rows(min_row=3, values_only=True):
            if row[0] and str(row[0]).strip() not in ('分店 ID', ''):
                channels.append({'id': str(row[0]).strip(), 'name': str(row[1]).strip() if row[1] else ''})
        we = wb['ID_employee']
        employees = []
        for row in we.iter_rows(min_row=3, values_only=True):
            if row[0] and str(row[0]).strip() not in ('在職員工 ID', ''):
                employees.append({'id': str(row[0]).strip(), 'name': str(row[1]).strip() if row[1] else ''})
        return products, channels, employees
    finally:
        try: os.unlink(tmp)
        except: pass

def group_orders(rows):
    from collections import defaultdict
    counts = defaultdict(int)
    for r in rows:
        counts[(r['code'], r['name'], r['color'], r['size'])] += 1
    return [{'code':k[0],'name':k[1],'color':k[2],'size':k[3],'qty':v} for k,v in counts.items()]

def run_matching(order_rows, products, store, staff):
    grouped = group_orders(order_rows)
    results, unmatched = [], []
    for row in grouped:
        prods = find_product(row['code'], row['name'], products)
        if not prods:
            unmatched.append({**row, 'issue':'找不到商品', 'avail':''}); continue
        match = find_variation(prods, row['color'], row['size'])
        if not match:
            avail = '、'.join(dict.fromkeys(p['variationCN'] for p in prods))
            unmatched.append({**row, 'issue':'找不到規格', 'avail':avail}); continue
        results.append({**match, 'qty': row['qty']})
    return results, unmatched, len(grouped)

def build_main_xlsx(results, store, staff):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template (dont move)'
    ws.append(['Purchase Order Handle*','Purchase Store ID*','Purchase Store Name',
               'Staff ID*','Staff Name','Supplier ID','Supplier Name',
               'Scheduled  Arrival Date','Other Fee','Custom ID','Remarks',
               'Product ID*','Product Name (Chinese)','Product Name (English)',
               'Variation ID','Variation (Chinese)','Variation (English)',
               'SKU','Purchase Quantity*','Cost'])
    ws.append(['進貨單辨識號碼*','進貨分店 ID*','進貨分店名稱','進貨人員 ID*','進貨人員名稱',
               '供應商 ID','供應商名稱','預訂到貨日期','其他費用','自訂單號','備註',
               '商品 ID*','商品名稱(繁體中文)','商品名稱(英文)','規格 ID',
               '規格名稱（繁體中文）','規格名稱（英文）','商品貨號','進貨數量*','成本'])
    for r in results:
        ws.append([1, store['id'], store['name'], staff['id'], staff['name'],
                   '','','','','','',
                   r['productId'], r['productName'], '',
                   r['variationId'], r['variationCN'], '',
                   r['sku'], r['qty'],
                   r['cost'] if r['cost'] != '' else ''])
    for col, w in enumerate([10,26,14,26,14,14,14,14,10,10,10,26,30,20,26,20,16,14,12,10], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    out = io.BytesIO(); wb.save(out); return out.getvalue()

def build_warn_xlsx(unmatched):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '待確認清單'
    headers = ['商品編號','商品名稱','進貨單規格','進貨數量','問題說明',
               'Shopline 現有規格（請對照填入）','✏️ 確認後的 Variation ID','✏️ 確認後的規格名稱']
    hf = PatternFill('solid', start_color='FFF2CC')
    rf = PatternFill('solid', start_color='FFF9E6')
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h); c.font = Font(bold=True); c.fill = hf
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    for i, u in enumerate(unmatched, 2):
        for col, v in enumerate([u['code'],u['name'],u['color'],u['qty'],u['issue'],u['avail'],'',''], 1):
            c = ws.cell(i, col, v)
            if col <= 6: c.fill = rf
    for col, w in enumerate([14,36,16,10,12,60,36,22], 1):
        ws.column_dimensions[ws.cell(1,col).column_letter].width = w
    out = io.BytesIO(); wb.save(out); return out.getvalue()

# ── Session store ─────────────────────────────────────────────────────────────
_sessions = {}

def new_session(data=None):
    sid = str(uuid.uuid4())[:8]
    _sessions[sid] = data or {}
    return sid

# ── Multipart parser ──────────────────────────────────────────────────────────
def parse_multipart(body, boundary):
    """Parse multipart form data. Returns dict of {name: value_bytes}"""
    result = {}
    if isinstance(boundary, str):
        boundary = boundary.encode()
    delimiter = b'--' + boundary
    parts = body.split(delimiter)
    for part in parts:
        part = part.strip(b'\r\n')
        if not part or part == b'--':
            continue
        # Split header and body
        for sep in [b'\r\n\r\n', b'\n\n']:
            if sep in part:
                head, _, content = part.partition(sep)
                break
        else:
            continue
        # Strip trailing CRLF from content
        while content.endswith(b'\r\n'):
            content = content[:-2]
        while content.endswith(b'\n'):
            content = content[:-1]
        # Parse field name from headers
        head_str = head.decode('utf-8', errors='ignore')
        name = ''
        for line in head_str.replace('\r\n', '\n').split('\n'):
            if 'Content-Disposition' in line:
                for seg in line.split(';'):
                    seg = seg.strip()
                    if seg.startswith('name='):
                        name = seg[5:].strip('"\'')
        if name:
            result[name] = content
    return result

# ── HTTP Handler ──────────────────────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args): pass

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == '/':
            body = HTML.encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        elif parsed.path == '/download':
            params = parse_qs(parsed.query)
            sid = params.get('session', [''])[0]
            dtype = params.get('type', ['main'])[0]
            data = _sessions.get(sid, {}).get(f'{dtype}_bytes')
            if not data:
                self.send_response(404); self.end_headers(); return
            fn = '進貨單_已填寫.xlsx' if dtype == 'main' else '待確認清單.xlsx'
            fn_encoded = fn.encode('utf-8').hex()
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f"attachment; filename*=UTF-8''{fn}")
            self.send_header('Content-Length', str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        else:
            self.send_response(404); self.end_headers()

    def do_POST(self):
        parsed = urlparse(self.path)
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length)
        if parsed.path == '/upload':
            self._handle_upload(body)
        elif parsed.path == '/process':
            self._handle_process(body)
        else:
            self.send_response(404); self.end_headers()

    def _json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _handle_upload(self, body):
        ct = self.headers.get('Content-Type', '')
        boundary = ''
        for part in ct.split(';'):
            part = part.strip()
            if part.startswith('boundary='):
                boundary = part[9:].strip()
        if not boundary:
            self._json({'error': '無效的上傳格式'}); return
        try:
            fields = parse_multipart(body, boundary)
            file_bytes = fields.get(b'file') or fields.get('file')
            ftype_bytes = fields.get(b'type') or fields.get('type')
            if file_bytes is None:
                self._json({'error': '未收到檔案'}); return
            ftype = ftype_bytes.decode('utf-8').strip() if ftype_bytes else 'order'
            if isinstance(file_bytes, memoryview):
                file_bytes = bytes(file_bytes)
            sid = new_session()
            if ftype == 'order':
                rows = parse_order_bytes(file_bytes)
                _sessions[sid]['order_rows'] = rows
                self._json({'session': sid, 'rows': len(rows)})
            else:
                products, channels, employees = parse_product_bytes(file_bytes)
                _sessions[sid]['products'] = products
                self._json({'session': sid, 'channels': channels,
                            'employees': employees, 'count': len(products)})
        except Exception as e:
            self._json({'error': f'檔案解析失敗：{str(e)}'})

    def _handle_process(self, body):
        try:
            req = json.loads(body)
            order_rows = _sessions.get(req.get('order_session',''), {}).get('order_rows')
            products   = _sessions.get(req.get('product_session',''), {}).get('products')
            store = req.get('store')
            staff = req.get('staff')
            if not order_rows: self._json({'error': '找不到進貨單，請重新上傳'}); return
            if not products:   self._json({'error': '找不到商品檔，請重新上傳'}); return
            results, unmatched, total = run_matching(order_rows, products, store, staff)
            main_bytes = build_main_xlsx(results, store, staff)
            warn_bytes = build_warn_xlsx(unmatched) if unmatched else None
            sid = new_session()
            _sessions[sid]['main_bytes'] = main_bytes
            if warn_bytes: _sessions[sid]['warn_bytes'] = warn_bytes
            self._json({
                'session': sid,
                'matched': len(results),
                'total': total,
                'unmatched': [{'code':u['code'],'name':u['name'],'color':u['color'],
                               'size':u['size'],'qty':u['qty'],'issue':u['issue'],
                               'avail':u['avail']} for u in unmatched]
            })
        except Exception as e:
            self._json({'error': f'處理失敗：{str(e)}'})

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    PORT = int(os.environ.get('PORT', '8080'))
    server = ThreadingHTTPServer(('0.0.0.0', PORT), Handler)
    print(f'Shopline 進貨單工具已啟動，Port: {PORT}')
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print('\n已關閉。')

if __name__ == '__main__':
    main()
